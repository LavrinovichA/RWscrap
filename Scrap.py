from bs4 import BeautifulSoup
import requests
import re
import openpyxl
import os
import time
from bs4.element import Tag

start_time = time.time()
os.system('cls')

print ('Разбираю файл ... \n')
file = 'Request.xlsx'
wb = openpyxl.load_workbook(file, data_only=True)
sheet = wb.active
wb_response = openpyxl.Workbook()
ws = wb_response.active
ws.append(['Parts number', 'Discount price', 'Price', 'Stock'])

os.system('cls')
print ('Отправляю запросы\n')
for i in range(2, sheet.max_row + 1):
#Ищем новую URL
    href_list = []
    sku = str(sheet.cell (row = i, column = 1).value)
    url = 'https://originalteile.mercedes-benz.de/search?search='+ sku
    html_text = requests.get(url).text
    soup = BeautifulSoup(html_text, 'html.parser')
    href_list = []
    for href in [div.a['href'] for div in soup.find_all('div',class_="product-image-wrapper")]:
        href_list.append(href)
        for ih in href_list:
            if sku in ih:
                new_url = ih
                break
            else:
                new_url = url
#Работаем с новой URL
    html_text = requests.get(new_url).text
    soup = BeautifulSoup(html_text, 'html.parser')
    try: 
        Default_price = re.findall(r'\d*\,\d+|\d+', str(soup.find('p', class_ = 'product-detail-price').text).replace('.', ''))[0]
    except:
        Default_price = 'Not Found'
    try:
        Stock = ((list(map(int, str(soup.find('select', class_ = 'custom-select product-detail-quantity-select').text).split())))[-1])
    except:
        Stock = 'Not Found'
    os.system('cls')
    print('Строка # ', i - 1)
    ws.append([sku, Default_price, '0', Stock])
    wb_response.save('Response.xlsx')
    
elapsed_time = time.time() - start_time
    
os.system('cls')
print ('\n', 'Выполнено: ', i - 1, ' запросов \n', 'За время:', elapsed_time, 'c \n', '\n', 'Все результаты сохранены в файл D:\DE_price_control\Response.xlsx', '\n')

